import openpyxl


class ExcelReader:

    def __init__(self, excel_path: str):
        self.path: str = excel_path
        self.wb = openpyxl.load_workbook(self.path)

    @staticmethod
    def _get_link_if_exists(cell) -> str | None:
        try:
            return cell.hyperlink.target
        except AttributeError:
            return None

    def get_data(self) -> dict:
        """
            Читает excel и возвращает словарь <номер маршрута по договору>: <id таблицы>.
        """
        data = dict()
        sheet = self.wb.active

        for i in range(1, sheet.max_row + 1):
            route_numb, link = sheet.cell(row=i, column=1).value, self._get_link_if_exists(sheet.cell(row=i, column=2))
            if link:
                if isinstance(route_numb, str):
                    if '(зима)' in route_numb:
                        route_numb = route_numb.replace('(зима)', '')
                    route_numb = route_numb.strip().upper()

                data[route_numb] = link.split('/')[5]

        return data



